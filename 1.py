import openpyxl
from openpyxl import Workbook
from datetime import datetime

def crear_archivo():
    try:
    
        libro = openpyxl.load_workbook('informe_gastos.xlsx')
        print("Archivo 'informe_gastos.xlsx' cargado.")
    except FileNotFoundError:
        
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Gastos"
        hoja.append(["Fecha", "Descripción", "Monto"]) 
        libro.save('informe_gastos.xlsx')
        print("Archivo 'informe_gastos.xlsx' creado.")
    return libro


def ingresar_gastos():
    gastos = []
    while True:
        
        while True:
            fecha_str = input("Ingrese la fecha del gasto (YYYY-MM-DD): ")
            try:
                fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                break
            except ValueError:
                print("Formato de fecha inválido. Intente nuevamente.")
        
        descripcion = input("Ingrese la descripción del gasto: ")
        
        
        while True:
            try:
                monto = float(input("Ingrese el monto del gasto: "))
                break
            except ValueError:
                print("Monto inválido. Debe ser un número. Intente nuevamente.")
        
        gastos.append({"fecha": fecha, "descripcion": descripcion, "monto": monto})
        
        continuar = input("¿Desea ingresar otro gasto? (s/n): ")
        if continuar.lower() != 's':
            break
    return gastos


def guardar_gastos(libro, gastos):
    hoja = libro["Gastos"]
    for gasto in gastos:
        hoja.append([gasto['fecha'], gasto['descripcion'], gasto['monto']])
    libro.save('informe_gastos.xlsx')
    print("Gastos guardados en 'informe_gastos.xlsx'.")


def mostrar_resumen(gastos, libro):
    if not gastos:
        print("No hay gastos para mostrar.")
        return
    
    total_gastos = sum(gasto['monto'] for gasto in gastos)
    gasto_mas_caro = max(gastos, key=lambda x: x['monto'])
    gasto_mas_barato = min(gastos, key=lambda x: x['monto'])
    
    print("\nResumen de gastos:")
    print(f"Número total de gastos: {len(gastos)}")
    print(f"Gasto más caro: {gasto_mas_caro['descripcion']} el {gasto_mas_caro['fecha']} por ${gasto_mas_caro['monto']:.2f}")
    print(f"Gasto más barato: {gasto_mas_barato['descripcion']} el {gasto_mas_barato['fecha']} por ${gasto_mas_barato['monto']:.2f}")
    print(f"Monto total de gastos: ${total_gastos:.2f}")
    
    
    hoja_resumen = libro.create_sheet("Resumen")
    hoja_resumen.append(["Resumen de Gastos"])
    hoja_resumen.append(["Número total de gastos", len(gastos)])
    hoja_resumen.append(["Gasto más caro", gasto_mas_caro['descripcion'], gasto_mas_caro['fecha'], f"${gasto_mas_caro['monto']:.2f}"])
    hoja_resumen.append(["Gasto más barato", gasto_mas_barato['descripcion'], gasto_mas_barato['fecha'], f"${gasto_mas_barato['monto']:.2f}"])
    hoja_resumen.append(["Monto total de gastos", f"${total_gastos:.2f}"])
    
    libro.save('informe_gastos.xlsx')
    print("Resumen guardado en 'informe_gastos.xlsx'.")


def main():
    libro = crear_archivo()  
    gastos = ingresar_gastos()  
    guardar_gastos(libro, gastos)  
    mostrar_resumen(gastos, libro)  

if __name__ == "__main__":
    main()
